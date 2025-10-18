import openpyxl
import logging
import requests
import socket
from logging_signals import log_signals
from PySide6.QtWidgets import QMainWindow
from PySide6.QtWidgets import *
from PySide6.QtGui import *
from PySide6.QtCore import *
from PySide6.QtMultimedia import *
from clock import Clock
from media_display_widget import MediaDisplayWidget
from config import FILE_PRAYER_TIMES_PATH, FILE_RAMADAN_PLAN, KALIMA_PATH, MEDIA_PLAYER_HEIGHT, WIDTH, HEIGHT, MEDIA_PLAYER_WIDTH, HEADER_HEIGHT, FOOTER_HEIGHT


class SubhanTvApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.logger = logging.getLogger(__name__)  
        self.logger.info("Initializing SubhanTvApp")
        log_signals.log_message.connect(self.handle_log_message)

        self.setWindowTitle("Subhan Moschee TV")
        #print(f"ss {self.size()}")
        #print(f"sss {self.screen().size()}")
        #self.setWindowFlags(Qt.Window | Qt.FramelessWindowHint)
        #self.showFullScreen()

# Data 
        # Überwachung der Gebetszeiten-Datei
        self.file_watcher = QFileSystemWatcher(self)
        self.file_watcher.addPath(FILE_PRAYER_TIMES_PATH)
        self.file_watcher.fileChanged.connect(self.reload_prayer_times)

        self.file_error_messages = []
        self.ramadan_plan = self.try_load_ramadan_plan(FILE_RAMADAN_PLAN)
        self.prayer_times = self.try_load_prayer_time_data(FILE_PRAYER_TIMES_PATH)
        self.last_checked_date = QDate.currentDate()
        #print(self.prayer_times.head())
        #print(self.prayer_times.dtypes)


# Create UI
        self.createUI()

        self.status_timer = QTimer()
        self.status_timer.timeout.connect(self.update_icons)
        self.status_timer.start(30000)  # alle 30 Sekunden

        self.update_icons()
# Timer für Datenauslesen
        '''
        self.update_prayer_timer = QTimer(self)
        self.update_prayer_timer.timeout.connect(self.check_for_updated_prayer_times)
        self.update_prayer_timer.start(60000)'''

    def reload_prayer_times(self):
        """Wird aufgerufen, wenn die Excel-Datei geändert wird"""
        logging.info("Gebetszeiten-Datei wurde aktualisiert. Lade neue Zeiten...")
        self.file_watcher.addPath(FILE_PRAYER_TIMES_PATH)
        try:
            self.prayer_times = self.try_load_prayer_time_data(FILE_PRAYER_TIMES_PATH)
            self.updatePrayerTimesUI()
        except Exception as e:
            logging.error(f"Fehler beim Laden der Gebetszeiten: {str(e)}")

    def try_load_ramadan_plan(self, file_path):
        try:
            ramadan_plan = []
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                date, sahar, iftar = row
                if date and sahar and iftar:
                    if isinstance(date, str):
                        date = QDate.fromString(date, "dd.MM.yyyy")
                    else:
                        date = QDate(date.year, date.month, date.day)

                    sahar = str(sahar)
                    iftar = str(iftar)

                    if QDate.currentDate() == date:
                        self.current_sahar_time = sahar
                        self.current_iftar_time = iftar

                    ramadan_plan.append({"Date": date, "Sahar": sahar, "Iftar": iftar})
            return ramadan_plan    
        except FileNotFoundError:
            error_message = f"File not found: {file_path}"
            logging.error(error_message)
        except Exception as e:
            error_message = f"An error occurred while loading the file: {file_path} - {str(e)}"
            logging.error(error_message)
        return []

    def try_load_prayer_time_data(self, file_path):
        try:
            prayer_times = []
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
                prayer, time = row
                if prayer and time:
                    time = time.strftime('%H:%M')

                    if isinstance(time, str):
                        qtime = QTime.fromString(time, "HH:mm")  # Falls Zeit als String gespeichert ist
                    else:
                        qtime = QTime(time.hour, time.minute)   # Falls Zeit als `datetime.time`-Objekt gespeichert ist
                    #print(f"Current Date {QDate.currentDate()}")    
                    if QDate.currentDate() < QDate(2025, 3, 31): # Wenn Ramadan noch nicht vorbei ist
                        qtime = QTime.fromString(time, "HH:mm")
                        if prayer == "Fajr":
                            #print(f"Sahar {self.current_sahar_time}")
                            qtime = QTime.fromString(self.current_sahar_time, "HH:mm")
                            qtime = qtime.addSecs(20 * 60)
                        elif prayer == "Maghrib":
                            #print(f"Iftar {self.current_iftar_time}")
                            qtime = QTime.fromString(self.current_iftar_time, "HH:mm")
                            qtime = qtime.addSecs(10 * 60)
                        time = qtime.toString("HH:mm")
                    prayer_times.append({"Prayer": prayer, "HH:MM": time})
                else:
                    continue

            if QDate.currentDate() < QDate(2025, 3, 31) and self.current_sahar_time and self.current_iftar_time:
                prayer_times.append({"Prayer": "Sahar", "HH:MM": self.current_sahar_time})
                prayer_times.append({"Prayer": "Iftar", "HH:MM": self.current_iftar_time})
            return prayer_times
        except FileNotFoundError:
            error_message = f"File not found: {file_path}"
            logging.error(error_message)
        except Exception as e:
            error_message = f"An error occurred while loading the file: {file_path} - {str(e)}"
            logging.error(error_message)
        return []

    def createUI(self):
        try:
            centralWidget = QWidget()
            self.setCentralWidget(centralWidget)        

# Set the HBox layout as main layout
            main_layout = QHBoxLayout(centralWidget)
            main_layout.setContentsMargins(0, 0, 0, 0)
            main_layout.setObjectName("mainLayout")
            centralWidget.setLayout(main_layout)

# Stacked widget
            self.stacked_widget = QStackedWidget()

            # Page 1
            self.createCombinedPage("Page-1")

            main_layout.addWidget(self.stacked_widget)

            self.timer = QTimer(self)
            self.timer.timeout.connect(self.updateTimeAndDate)
            self.timer.start(1000)

        except Exception as e:
            error_message = f"Failed to initialize the UI: {str(e)}"
            logging.error(error_message)
            #print(error_message) 

    def updateTimeAndDate(self):
        try:
            locale = QLocale(QLocale.German, QLocale.Germany)
            currentTime = QTime.currentTime().toString("HH:mm:ss")
            self.digital_clock.setText(currentTime)

            currentDate = locale.toString(QDate.currentDate(), "dddd, dd.MM.yyyy")
            if self.date_label.text() != currentDate:
                self.date_label.setText(currentDate)
            
            # Datum prüfen und ggf. Sahar/Iftar aktualisieren
            if QDate.currentDate() != self.last_checked_date:
                self.last_checked_date = QDate.currentDate()
                self.logger.info("Neuer Tag erkannt aktualisiere Sahar und Iftar Zeiten.")
                self.updatePrayerTimesUI()

        except Exception as e:
            error_message = f"Error updating time and date: {str(e)}"
            logging.error(error_message)
            #print(error_message)    
        
        self.analog_clock.update()

    def createCombinedPage(self, pageName):
        combined_widget = QWidget()
        combined_layout = QHBoxLayout(combined_widget)
        combined_layout.setContentsMargins(0, 0, 0, 0)
        if pageName == "Page-1":
            combined_layout.addWidget(self.createPage1())
        elif pageName == "Page-2":
            combined_layout.addWidget(self.createPage1())
        elif pageName == "Page-3":
            combined_layout.addWidget(self.createPage1())
        self.stacked_widget.addWidget(combined_widget)

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Escape: 
            self.showMaximized()  
        elif event.key() == Qt.Key_F: 
            #if self.isFullScreen():
                #self.showMaximized() 
            #else:
                self.showFullScreen()  
        elif event.key() == Qt.Key_S:
            self.close()
        else:
            super().keyPressEvent(event)
    
    def createPage1(self):
# Page central Layout
        vbox_widget = QWidget()
        vbox_layout = QVBoxLayout()
        #vbox_widget.setFixedHeight(30)
        #vbox_widget.setObjectName("header_layout_widgets")
        vbox_widget.setLayout(vbox_layout)
        vbox_widget.setFixedWidth(WIDTH)
        vbox_widget.setFixedHeight(HEIGHT)

# Header Layout
        hbox_header_widget = QWidget()
        hbox_header_layout = QHBoxLayout()
        hbox_header_layout.setContentsMargins(0, 0, 0, 0)
        hbox_header_widget.setLayout(hbox_header_layout)
        hbox_header_widget.setObjectName("header_layout_widgets")
        hbox_header_widget.setFixedHeight(HEADER_HEIGHT)
        hbox_header_widget.setFixedWidth(WIDTH)
        vbox_layout.addWidget(hbox_header_widget)
        
        self.kalima_label = QLabel()
        kalima_pixmap = QPixmap(KALIMA_PATH)  # Ersetze mit dem tatsächlichen Pfad zum Bild
        fixed_width = 350 
        kalima_pixmap = kalima_pixmap.scaled(fixed_width, fixed_width, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        self.kalima_label.setPixmap(kalima_pixmap)
        self.kalima_label.setAlignment(Qt.AlignCenter)  # Zentriere das Bild, wenn gewünscht
        hbox_header_layout.addWidget(self.kalima_label)
    
        self.kalma_translation = QLabel("Niemand ist anbetungswürdig außer Gott, Muhammad (saw) ist der Gesandte Allahs.")
        self.kalma_translation.setProperty("class", "header_text")
        self.kalma_translation.setAlignment(Qt.AlignCenter)
        hbox_header_layout.addWidget(self.kalma_translation)

        self.kalima_label_effect = QGraphicsOpacityEffect(self.kalima_label)
        self.kalima_label.setGraphicsEffect(self.kalima_label_effect)
        self.kalima_label_effect.setOpacity(1.0)  # Starte sichtbar

        self.kalma_translation_effect = QGraphicsOpacityEffect(self.kalma_translation)
        self.kalma_translation.setGraphicsEffect(self.kalma_translation_effect)
        self.kalma_translation_effect.setOpacity(0.0)  # Starte unsichtbar

        self.kalima_label_animation = QPropertyAnimation(self.kalima_label_effect, b"opacity")
        self.kalima_label_animation.setDuration(1000)  # Dauer der Animation in Millisekunden
        self.kalima_label_animation.setStartValue(1.0)
        self.kalima_label_animation.setEndValue(0.0)

        self.kalma_translation_animation = QPropertyAnimation(self.kalma_translation_effect, b"opacity")
        self.kalma_translation_animation.setDuration(1000)
        self.kalma_translation_animation.setStartValue(0.0)
        self.kalma_translation_animation.setEndValue(1.0)

        self.kalma_translation.setVisible(False)

        self.header_timer = QTimer(self)
        self.header_timer.timeout.connect(self.toggleHeaderWidgets)
        self.header_timer.start(10000)

# Main Layout 1
        hbox_middle_widget = QWidget()
        hbox_middle_widget.setProperty("class", "mediaPlayer")
        hbox_middle_layout = QHBoxLayout()
        hbox_middle_layout.setContentsMargins(0, 0, 0, 0)
        hbox_middle_widget.setLayout(hbox_middle_layout)

        mediaPlayer_widget = MediaDisplayWidget()
        mediaPlayer_widget.setFixedHeight(MEDIA_PLAYER_HEIGHT)
        mediaPlayer_widget.setFixedWidth(MEDIA_PLAYER_WIDTH)

        hbox_middle_layout.addWidget(mediaPlayer_widget)

#Main Layout 2
        sideBar_widget = QWidget()
        sideBar_widget.setProperty("class", "clock_widgets")
        sideBar_widget.setFixedHeight(MEDIA_PLAYER_HEIGHT)
        sideBar_widget.setFixedWidth(WIDTH-MEDIA_PLAYER_WIDTH)
        
        self.sideBar_layout = QVBoxLayout()
        self.sideBar_layout.addStretch()

        self.analog_clock = Clock(self)
        self.sideBar_layout.addWidget(self.analog_clock)
        
        self.digital_clock = QLabel("00:00:00")
        self.digital_clock.setAlignment(Qt.AlignCenter)
        self.digital_clock.setObjectName("digital_clock")
        self.sideBar_layout.addWidget(self.digital_clock)

        self.date_label = QLabel()
        self.date_label.setAlignment(Qt.AlignCenter)
        self.date_label.setObjectName("date")
        self.sideBar_layout.addWidget(self.date_label)

        gebetszeiten = QLabel("Gebetszeiten")
        gebetszeiten.setObjectName("gebetszeiten_heading")
        gebetszeiten.setAlignment(Qt.AlignCenter)
        self.sideBar_layout.addWidget(gebetszeiten)

        self.addPrayerTimes()

        status_layout = QHBoxLayout()
        status_layout.setAlignment(Qt.AlignRight)
        status_layout.setContentsMargins(10, 5, 10, 5)
        status_layout.setSpacing(12)

        # Beispielhafte Icons – du kannst Pfade zu PNGs verwenden
        self.internet_icon = QLabel()
        self.internet_icon.setPixmap(QPixmap("assets/wifi.png").scaled(24, 24, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        status_layout.addWidget(self.internet_icon)

        self.warning_icon = QLabel()
        self.warning_icon.setPixmap(QPixmap("assets/warning.png").scaled(24, 24, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        status_layout.addWidget(self.warning_icon)
        self.warning_icon.setVisible(False)

        self.error_icon = QLabel()
        self.error_icon.setPixmap(QPixmap("assets/error.png").scaled(24, 24, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        status_layout.addWidget(self.error_icon)
        self.error_icon.setVisible(False)

        self.tunnel_icon = QLabel()
        self.tunnel_icon.setPixmap(QPixmap("assets/tunnel.png").scaled(24, 24, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        status_layout.addWidget(self.tunnel_icon)

        status_widget = QWidget()
        status_widget.setLayout(status_layout)
        status_widget.setProperty("class", "status_bar")

        self.sideBar_layout.addWidget(status_widget)       



        #self.sideBar_layout.addStretch()
        self.sideBar_layout.setContentsMargins(0, 0, 0, 0)
        sideBar_widget.setLayout(self.sideBar_layout)
        hbox_middle_layout.addWidget(sideBar_widget)
        
        hbox_middle_layout.addWidget(mediaPlayer_widget, 8) 
        hbox_middle_layout.addWidget(sideBar_widget, 2)
        hbox_middle_layout.setContentsMargins(0, 0, 0, 0)
        hbox_middle_layout.setSpacing(0)

        vbox_layout.setContentsMargins(0, 0, 0, 0)
        vbox_layout.addWidget(hbox_middle_widget)
        
# Footer Layout
        footer_widget = QWidget()
        footer_widget.setFixedHeight(FOOTER_HEIGHT)
        footer_widget.setFixedWidth(WIDTH)
        footer_widget.setObjectName("footer_layout_widgets")
        footer_layout = QHBoxLayout()
        footer_layout.setContentsMargins(0, 0, 0, 0)
        footer_widget.setLayout(footer_layout)

        jamaat_info_name1 = QLabel("Ahmadiyya Muslim Jamaat Mörfelden-Walldorf")
        jamaat_info_name1.setProperty("class", "footer_text")
        jamaat_info_name1.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        footer_layout.addWidget(jamaat_info_name1)

        jamaat_info_name2 = QLabel("Liebe Für Alle, Hass Für Keinen.")
        jamaat_info_name2.setProperty("class", "footer_text")
        jamaat_info_name2.setAlignment(Qt.AlignCenter)
        footer_layout.addWidget(jamaat_info_name2)

        jamaat_info_name3 = QLabel("www.ahmadiyya.de")
        jamaat_info_name3.setProperty("class", "footer_text")
        jamaat_info_name3.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        footer_layout.addWidget(jamaat_info_name3)

        vbox_layout.setSpacing(0)
        vbox_layout.addWidget(footer_widget)
        vbox_layout.addStretch(1)
        return vbox_widget
    
    def has_internet_connection(self, url="https://www.google.com", timeout=3):
        try:
            requests.get(url, timeout=timeout)
            return True
        except requests.RequestException:
            return False

    def is_tunnel_active(self, host="localhost", port=22, timeout=2):
        try:
            with socket.create_connection((host, port), timeout=timeout):
                return True
        except (socket.timeout, ConnectionRefusedError, OSError):
            return False

    def update_icons(self):
        if self.has_internet_connection():
            internet_icon_path = "assets/wifi.png"
            #tooltip = "Internetverbindung: OK"

            if self.is_tunnel_active():  
                tunnel_icon_path = "assets/tunnel.png"
            else:
                tunnel_icon_path = "assets/no-tunnel.png"
        else:
            internet_icon_path = "assets/no-wifi.png"
            #tooltip = "Keine Internetverbindung"
            tunnel_icon_path = "assets/no-tunnel.png"

        self.internet_icon.setPixmap(QPixmap(internet_icon_path).scaled(24, 24, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        self.tunnel_icon.setPixmap(QPixmap(tunnel_icon_path).scaled(24, 24, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        #self.internet_icon.setToolTip(tooltip)

    def handle_log_message(self, level, message):
        if level in ["ERROR", "CRITICAL"]:
            self.error_icon.setVisible(True)
            self.error_icon.setToolTip(message)
            QTimer.singleShot(60000, lambda: self.error_icon.setVisible(False))
        elif level == "WARNING":
            self.warning_icon.setVisible(True)
            self.warning_icon.setToolTip(message)
            QTimer.singleShot(60000, lambda: self.warning_icon.setVisible(False))



    def addPrayerTimes(self):
        if self.prayer_times:
            self.prayer_widget = QWidget()
            prayer_gridLayout = QGridLayout()
            prayer_gridLayout.setSpacing(0)
            prayer_gridLayout.setContentsMargins(0, 0, 0, 0)

            index = 0
# Korrekte Grid hier erstellen
            for row in self.prayer_times:        
                prayer = row['Prayer']
                time = row['HH:MM']

                if prayer == "Sahar":
                    separator = QFrame()
                    separator.setFrameShape(QFrame.HLine)  # Horizontale Linie
                    separator.setProperty("class", "separator")
                    prayer_gridLayout.addWidget(separator, index, 0, 1, 2)  # Linie einfügen
                    index += 1
                
                self.prayer_label = QLabel(prayer)
                self.prayer_label.setProperty("class", "salat_timings")
                prayer_gridLayout.addWidget(self.prayer_label, index, 0)

                time_label = QLabel(str(time))
                time_label.setProperty("class", "salat_timings")
                prayer_gridLayout.addWidget(time_label, index, 1)
            
                index += 1

            self.prayer_widget.setLayout(prayer_gridLayout)
            self.sideBar_layout.addWidget(self.prayer_widget)

    def updatePrayerTimesUI(self):
        """ Ersetzt das Widget mit den neuen Gebetszeiten """
        try:
            # Entferne das alte Widget
            if self.prayer_widget:
                self.prayer_widget.setParent(None)  # Entfernt es aus dem Layout

            # Erstelle ein neues Widget mit aktualisierten Gebetszeiten
            self.addPrayerTimes()
            
            logging.info("Gebetszeiten in der UI erfolgreich aktualisiert.")
        except Exception as e:
            logging.error(f"Fehler beim Aktualisieren der Gebetszeiten: {str(e)}")


    def toggleHeaderWidgets(self):
        # Entscheide, welche Animation in welche Richtung läuft
        if self.kalima_label_effect.opacity() == 1.0:
            self.kalima_label_animation.setDirection(QPropertyAnimation.Forward)
            self.kalima_label.setVisible(False)
            self.kalma_translation_animation.setDirection(QPropertyAnimation.Forward)
            self.kalma_translation.setVisible(True)
        else:
            self.kalima_label_animation.setDirection(QPropertyAnimation.Backward)
            self.kalima_label.setVisible(True)
            self.kalma_translation_animation.setDirection(QPropertyAnimation.Backward)
            self.kalma_translation.setVisible(False)

        # Starte die Animationen
        self.kalima_label_animation.start()
        self.kalma_translation_animation.start()

